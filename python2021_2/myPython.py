import os   # カレントディレクトリ変更で使用
import shutil   # ディレクトリコピーに使用
import datetime  # 保存時の日付情報で使用
import glob  # エクセル一覧の取得で使用

import openpyxl  # エクセル操作。要「pip install openpyxl」。ちなみに端末内にエクセルがインストールされていなくても動く。

from concurrent.futures import ThreadPoolExecutor

ORIGINAL_EXCEL_DIRECTORY = ""  # インプットディレクトリ。元ネタ。作業用ディレクトリ作成時に、最初の1回だけ参照する。
WORK_DIRECTORY = ""  # 作業用ディレクトリ。bakや作成中のディレクトリを削除。ここのエクセルを読み込んで、markdownを生成する。
OUTPUT_DIRECTORY = ""  # アウトプットディレクトリ。ここにmdを生成する。
IS_SKIP_INIT = False    # 初期化処理（init()）を実行するかどうか。スキップする場合（True）、work ディレクトリ とかを毎回やらない。めんどくさい人用。
CSS_TEXT = ('/* 画像の区切りが見づらいとの指摘に対して黒い枠線を付与 */\n'  # style.css の中身。もうこういうファイルを用意して配布するタイプの方がいい気がしてきた。最初は数行だったんだよ。
            'img { border: 1px black solid;}\n'
            '\n'
            '/* 見出しが分かりづらいとの指摘に対して各種文字飾りを付与*/\n'
            'h1 { font-weight:bold;}\n'
            'h2 { color: midnightblue;}\n'
            'h3 { background-color: darkblue; font-weight:bold; color: white}\n'
            'h4 { border: 5px lightgrey double;}\n'
            '\n'
            'table {border: 1px solid #e0e0e0;}\n'
            'th {background: #f0f0f0; border-left: 1px solid #e0e0e0;}\n'
            'td {border-left: 1px solid #e0e0e0;}\n'
            'hr {color: midnightblue;}\n'
            '\n'
            'footer {\n'
            '    background: midnightblue;\n'
            '    color: white;\n'
            '    padding-left: 30px;\n'
            '    padding-right: 30px;\n'
            '    box-sizing: border-box;\n'
            '}'
            )


class ScreenDesignData:
    def __init__(self):
        self.title = ''  # タイトル
        self.history = []   # 履歴
        self.layout = {}    # レイアウト : 概要となんか
        self.layoutItems = []  # 画面項目
        self.inputItems = {}  # 入力項目
        self.events = []    # イベント一覧
        self.inputCheck = []    # 入力チェック
        self.businessCheck = []  # 業務チェック
        self.functions = {}  # 機能設計群

        # Markdown定型文
        self.markdownTemplate = ('<link rel="stylesheet" type="text/css" href="style.css">\n'
                                 '\n'
                                 '# イトーヨーカドーネットスーパー<br><br>@specTitle 画面設計書\n'
                                 '\n'
                                 '株式会社ビッグツリーテクノロジー＆コンサルティング\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 改訂履歴\n'
                                 '\n'
                                 '@specHistory'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 概要\n'
                                 '\n'
                                 '```Overview\n'
                                 '@specOverview\n'
                                 '```\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## レイアウト\n'
                                 '\n'
                                 '  - 画面タイトル1  \n'
                                 '    ![画面1](img\\1.jpg)\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 画面項目\n'
                                 '\n'
                                 '@specLayoutItems'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 入力項目\n'
                                 '\n'
                                 '@specInputItems'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## イベント一覧\n'
                                 '\n'
                                 '@specEvents'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 各種チェック\n'
                                 '\n'
                                 '### 入力チェック  \n'
                                 '\n'
                                 '@specInputCheck'
                                 '\n'
                                 '### 業務チェック  \n'
                                 '\n'
                                 '@specBusinessCheck'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 機能\n'
                                 '\n'
                                 '@functions'
                                 '\n'
                                 '<footer>以上</footer>'
                                 )

        self.markdownTemplate_functions = ('### @id : @name\n'
                                           '\n'
                                           '#### 入力\n'
                                           '\n'
                                           '@input\n'
                                           '\n'
                                           '#### 出力\n'
                                           '\n'
                                           '@output\n'
                                           '\n'
                                           '#### 処理内容\n'
                                           '\n'
                                           '```\n'
                                           '@processDetail'
                                           '```\n'
                                           '\n'
                                           )

    def generate_markdown(self):

        # マークダウン用の文字列を生成
        output_str = (self.markdownTemplate
                      .replace('@specTitle', self.title)
                      .replace('@specHistory', array_to_markdown_table(self.history, '改訂履歴'))
                      .replace('@specOverview', self.layout['Overview'])
                      .replace('@specLayoutItems', array_to_markdown_table(self.layoutItems, '画面項目'))
                      .replace('@specInputItems', self.generate_input_items())
                      .replace('@specEvents', array_to_markdown_table(self.events, 'イベント一覧'))
                      .replace('@specInputCheck', array_to_markdown_table(self.inputCheck, '入力チェック'))
                      .replace('@specBusinessCheck', array_to_markdown_table(self.businessCheck, '業務チェック'))
                      .replace('@functions', self.generate_functions())
                      )

        return output_str

    def generate_input_items(self):

        # ガード節
        if self.inputItems is None or len(self.inputItems) == 0:
            return '- 定義なし\n'

        return_str = ''
        sorted_items = sorted(self.inputItems.items(), key=lambda x: x[0])
        for k, v in sorted_items:    # キーでソートしながら中身を取り出す。sheetNamesはシートの並び順だが、dictの順序が保証されてるか怪しかったので。

            if v['title'] != '入力項目':    # 入力項目が複数存在したら、一つ小さい見出しでシート名を付与する。
                return_str = return_str + '### ' + v['title'] + '\n\n'

            return_str = return_str + (array_to_markdown_table(v['data'], '入力項目')) + '\n'

        return return_str

    def generate_functions(self):
        # ガード節
        if self.functions is None or len(self.functions) == 0:
            return '- 定義なし\n'

        return_str = ''
        sorted_functions = sorted(self.functions.items(), key=lambda x: x[0])
        for k, v in sorted_functions:    # キーでソートしながら中身を取り出す。sheetNamesはシートの並び順だが、dictの順序が保証されてるか怪しかったので。
            return_str = return_str + (self.markdownTemplate_functions
                                       .replace('@id', str(v['id'] or ''))
                                       .replace('@name', v['name'])
                                       .replace('@input', '- ' + v['input'].replace('\n', '\n- '))
                                       .replace('@output', '- ' + v['output'].replace('\n', '\n- '))
                                       .replace('@processDetail', v['processDetail'])
                                       )
        return return_str


class ReportData:   # 帳票設計書クラス

    def __init__(self):
        self.title = ""  # タイトル
        self.history = []   # 履歴
        self.layout = {}    # レイアウト : 概要となんか
        self.reportItems = []  # 帳票項目
        self.events = []    # イベント一覧

        # Markdown定型文
        self.markdownTemplate = ('<link rel="stylesheet" type="text/css" href="style.css">\n'
                                 '\n'
                                 '# イトーヨーカドーネットスーパー<br><br>@specTitle 帳票設計書\n'
                                 '\n'
                                 '株式会社ビッグツリーテクノロジー＆コンサルティング\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 改訂履歴\n'
                                 '\n'
                                 '@specHistory'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 概要\n'
                                 '\n'
                                 '```Overview\n'
                                 '@specOverview\n'
                                 '```\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## レイアウト\n'
                                 '\n'
                                 '  - 画面タイトル1  \n'
                                 '    ![画面1](img\\1.jpg)\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 帳票項目\n'
                                 '\n'
                                 '@specReportItems'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## イベント一覧\n'
                                 '\n'
                                 '@specEvents'
                                 '\n'
                                 '<footer>以上</footer>'
                                 )

    def generate_markdown(self):

        # マークダウン用の文字列を生成
        output_str = (self.markdownTemplate
                      .replace('@specTitle', self.title)
                      .replace('@specHistory', array_to_markdown_table(self.history, '改訂履歴'))
                      .replace('@specOverview', self.layout['Overview'])
                      .replace('@specReportItems', array_to_markdown_table(self.reportItems, '帳票項目'))
                      .replace('@specEvents', array_to_markdown_table(self.events, 'イベント一覧'))
                      )

        return output_str


class MailData:   # メール設計書クラス

    def __init__(self):
        self.title = ""  # タイトル
        self.history = []   # 履歴
        self.mailTemplate = {}    # メールテンプレート
        self.mailItems = []  # メール項目
        self.sample = []    # サンプル

        # Markdown定型文
        self.markdownTemplate = ('<link rel="stylesheet" type="text/css" href="style.css">\n'
                                 '\n'
                                 '# イトーヨーカドーネットスーパー<br><br>@specTitle メール設計書\n'
                                 '\n'
                                 '株式会社ビッグツリーテクノロジー＆コンサルティング\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## 改訂履歴\n'
                                 '\n'
                                 '@specHistory'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## メールテンプレート\n'
                                 '\n'
                                 '### 概要\n'
                                 '\n'
                                 '```Overview\n'
                                 '@specOverview\n'
                                 '```\n'
                                 '\n'
                                 '### メールテンプレート\n'
                                 '\n'
                                 '※ ***[＠＠＠＠]***・・・可変項目を斜体・太字で記載。可変項目の内容はメール項目参照。\n'
                                 '@specMailTemplate\n'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## メール項目\n'
                                 '\n'
                                 '@specMailItems'
                                 '\n'
                                 '------------------------------------------------------------------------------------------\n'
                                 '\n'
                                 '## サンプル\n'
                                 '\n'
                                 '```Sample\n'
                                 '@specSample'
                                 '```\n'
                                 '\n'
                                 '<footer>以上</footer>'
                                 )

    def generate_markdown(self):

        # マークダウン用の文字列を生成
        output_str = (self.markdownTemplate
                      .replace('@specTitle', self.title)
                      .replace('@specHistory', array_to_markdown_table(self.history, '改訂履歴'))
                      .replace('@specOverview', self.mailTemplate['Overview'])
                      .replace('@specMailTemplate', array_to_markdown_table(self.mailTemplate['MailTemplate'], 'メールテンプレート'))
                      .replace('@specMailItems', array_to_markdown_table(self.mailItems, 'メール項目'))
                      .replace('@specSample', self.generate_sample())
                      )

        return output_str

    def generate_sample(self):

        # ガード節
        if self.sample is None or len(self.sample) == 0:
            return '- 定義なし\n'

        # 配列が突っ込んであるので、Markdown用の1文に整形する
        return '\n'.join(self.sample)


def array_to_markdown_table(array, sheet_title):    # 配列をマークダウンのテーブル形式に変換

    # ガード節 : もし配列に項目レコードしかない場合（=データレコードが無い場合）、表が項目だけあってもしょうがないので、「- 定義なし」とかを返す
    if array is None or len(array) == 1:
        return '- 定義なし\n'

    s = ''
    s = '| ' + ' | '.join(array[0]).replace('\n', '<br>') + ' |\n'    # 1配列目は項目行。これはかならず存在する。|項番|aaa|bbb|ccc|ddd|とかの。
    s = s + '|:--' * len(array[0]) + '|\n'    # 間に表のあれ（アライメント？）を入れる。|:--|:--|:--|

    for arr in array[1:]:   # ここからデータ部整形

        # 結合に向けての準備
        for col in range(len(arr)):
            # 先頭の方はガード節やシート限定の特殊処理を記載
            if arr[col] is None:    # ガード節
                arr[col] = ' '    # None（値が入っていなかったセル）は「 」（半角スペース）を設定。マークダウンの表として「 」が必要なので。

            elif sheet_title == '画面項目' and col == 11:  # 画面設計書の画面項目の導出元（12列目） は 改行コードがあれば改行を2つ重ねる。そういうもん。
                arr[col] = arr[col].replace('\n', '<br><br>')  # str 型なら 改行コード（\n）の存在に気をつけて、基本はそのまま採用。

            elif sheet_title == '帳票項目' and col == 10:   # 帳票設計書の帳票項目の導出元（１１列目）は 改行コードがあれば改行を2つ重ねる。そういうもん。
                arr[col] = arr[col].replace('\n', '<br><br>')  # str 型なら 改行コード（\n）の存在に気をつけて、基本はそのまま採用。

            elif sheet_title == '改訂履歴' and col == 1:  # シート「改訂履歴」専用処理。
                # なお、まれに各セルが 数値や日付 + フォーマット ではなく文字列としてそのまま書かれている状況もある。それはもうわざとやっているとみなし、先頭の分岐でそのまま採用している。
                arr[col] = str('{:.2f}'.format(round(arr[col], 2)))  # シート「改訂履歴」の2列目はフォーマット指定の版数。値と実態が異なるので変換する。
                # TODO 端数処理:切り捨て
            elif sheet_title == '改訂履歴' and col == 2:
                arr[col] = f'{arr[col]:%Y/%m/%d}'   # 日付型をフォーマットする。「2021/01/01」形式

            # この辺から一般処理
            elif isinstance(arr[col], str):
                arr[col] = arr[col].replace('\n', '<br>')  # str 型なら 改行コード（\n）の存在に気をつけて、基本はそのまま採用。

            else:
                arr[col] = str(arr[col]).replace('\n', '<br>')  # なんかわからないものはすべてstr型に変更する。改行コード（\n）の存在に気をつけて、基本はそのまま採用。

        s = s + '| ' + ' | '.join(arr) + ' |\n'

    if sheet_title in ['入力チェック', '業務チェック']:
        s = s + (
            '\n'
            '<div class="annotation">*メッセージ表示位置画面 項目項番が「－」ハイフンの場合は共通のメッセージエリアに表示する。</div>\n'
        )

    return s


def init():

    print('\n★★ 初期化処理 - start')

    # もし「work」があるなら削除
    try:
        shutil.rmtree('work')
    except FileNotFoundError:
        pass

    # ディレクトリをまるごと別ディレクトリとして同階層にコピー。名前は workDir のとおり。
    shutil.copytree(ORIGINAL_EXCEL_DIRECTORY, WORK_DIRECTORY)

    # お掃除 : 「bak」ディレクトリは削除。
    file_list = glob.glob(WORK_DIRECTORY + '\\**\\bak', recursive=True)
    for file in file_list:
        print(file)
        try:
            shutil.rmtree(file)
            continue

        except FileNotFoundError:
            continue

    # お掃除 : 「作成中」ディレクトリは削除。つーかこういうのもうGitとかで管理して、完成したやつだけコミットしろや。履歴管理は履歴管理サービスに任せぇ！
    file_list = glob.glob(WORK_DIRECTORY + '\\**\\*作成中', recursive=True)
    for file in file_list:
        print(file)
        try:
            shutil.rmtree(file)
            continue

        except FileNotFoundError:
            continue

    # お掃除 : 「bk」ディレクトリは削除。
    file_list = glob.glob(WORK_DIRECTORY + '\\**\\bk', recursive=True)
    for file in file_list:
        print(file)
        try:
            shutil.rmtree(file)
            continue

        except FileNotFoundError:
            continue

    # 「.xlsx」をzipにして別名保存 : これやるかどうか微妙

    print('\n★★ 初期化処理 - end')


def read_sheet(wb, title):

    ws = None
    try:
        ws = wb[title]
    except KeyError:
        print('■ readError - ' + title)
        return None

    # まず表の開始位置（項番のセル）と表の右端を探索する。開始位置から行カーソルをwhileループする。右端まではforループで、それぞれ反復処理する。
    # 行で探索する対象の列も定義する。変数はtargetcolとする。
    r = targetcol = maxcol = 1  # それぞれの行列の変数。内容は↑の通り。

    list = []

    # 表の開始位置を探索
    while ws.cell(r, 1).value != '項番':  # 「項番」は必ず列「A」に存在するので、それを目指して行カーソル（r）を進める。
        r = r + 1

    # 表の右端の列番号を探索。ただし改訂履歴は2段になってるせいでとりあえずあとで固定値でいいや。
    while ws.cell(r, maxcol).value is not None:
        maxcol = maxcol + 1

    if ws.title == '改訂履歴':
        # 改訂履歴だけ項目が2段になってるから2つずらすし、列は「改定箇所」のセルを対象とする。
        r = r + 2
        targetcol = 5

        # ついでに項目ももうこっちで作っちゃう。他の表はセルを読んで自動取得+生成だけど、こいつめんどくさい。
        list.append(['項番', '版数', '更新日付', '更新者', '改定箇所', '改定内容', '改定理由'])

    while ws.cell(r, targetcol).value is not None:  # 行 ： 空っぽのセルが出てくるまでループする。
        list.append([ws.cell(r, n).value for n in range(1, maxcol)])  # 列「A」から「maxcol」までを1行分の列ループ。ちなみに空のセルは「None」が入る。
        r = r + 1

        # もしも「特定の列はスキップ」という要求があれば、for 文を分解して以下のようにする。
        # for n in range(1, maxcol):
        #     if ws.cell(r, n).value == 'aaa':
        #         continue
        #     array.append(ws.cell(r, n).value)
        # r = r + 1

    return list


def read_sheet_layout(ws):

    # レイアウト読み込み - ここから
    row = col = 1   # 行カーソルと列カーソルをセルの「A1」で初期化。ここから下に探索する。

    # 概要欄探索
    while ws.cell(row, col).value != '概要':
        row = row + 1

    dict = {'Overview': ws.cell(row + 1, col).value}    # 概要欄ゲット
    # レイアウト読み込み - ここまで

    return dict


def read_sheets_for_screen_design(title, wb):

    d = ScreenDesignData()  # インスタンス化

    d.title = title  # タイトル設定

    d.history = read_sheet(wb, '改訂履歴')  # 改訂履歴

    d.layout = read_sheet_layout(wb['レイアウト'])    # レイアウト

    d.layoutItems = read_sheet(wb, '画面項目')    # 画面項目読み込み

    d.events = read_sheet(wb, 'イベント一覧')    # イベント一覧読み込み

    d.inputCheck = read_sheet(wb, '入力チェック')    # 入力チェック読み込み

    d.businessCheck = read_sheet(wb, '業務チェック')    # 業務チェック読み込み

    def read_sheet_input_items(wb, s):    # 入力項目 シートの読み込み。1シートずつ

        # シート名を設定して、あとは行列を通常通り取得
        data = {'title': s}
        data['data'] = read_sheet(wb, s)

        return data   # 色々詰まったdataを返却

    def read_sheet_function(ws):    # 【機能】シートの読み込み。1シートずつ

        # 変数とか準備
        row = col = maxrow = maxcol = 1
        data = {}
        while ws.cell(maxrow, 1).value != '★END':  # end位置（下端）を特定
            maxrow = maxrow + 1
            if maxrow > 500:   # これぐらい到達しなかったらもうそこまででいいよ
                break

        while ws.cell(1, maxcol).value != '★END':  # end位置（右端）を特定
            maxcol = maxcol + 1
            if maxcol > 100:    # これぐらい到達しなかったらもうそこまででいいよ
                break

        # 機能IDと機能名
        while ws.cell(row, 1).value != '機能ID':    # 機能名と機能IDが存在する行まで行カーソルを移動
            row = row + 1

        data['id'] = ws.cell(row + 1, 1).value or ''    # 機能IDを取得

        while ws.cell(row, col).value != '機能名':  # 機能名が存在する列まで列カーソルを移動
            col = col + 1

        data['name'] = ws.cell(row + 1, col).value or ''  # 機能名を取得

        # 入力と出力
        col = 1
        while ws.cell(row, 1).value != '入力':    # 行カーソル移動
            row = row + 1

        data['input'] = ws.cell(row + 1, 1).value or ''    # 入力の内容を取得

        while ws.cell(row, col).value != '出力':  # 列カーソル移動
            col = col + 1

        data['output'] = ws.cell(row + 1, col).value or ''  # 出力の内容を取得

        # 処理内容
        col = 1
        while ws.cell(row, 1).value != '繰返':    # 行カーソル移動
            row = row + 1

        values = []
        for rowdata in ws.iter_rows(min_row=row + 1, max_row=maxrow, min_col=3, max_col=maxcol):  # 行の内容を指定の範囲で 1セルずつ取得。数値とかが入ってても困るので文字列にキャストする。
            values.append([str(cell.value) if cell.value is not None else ' ' for cell in rowdata])

        process_detail = ''
        for v in values:
            process_detail = process_detail + ''.join(v).rstrip() + '\n'  # 1行としておかしくない感じにトリムしたり改行コードであれしたりと、全体で1文になるように操作

        data['processDetail'] = process_detail

        return data   # 色々詰まったdataを返却

    # 【機能】xxx、入力項目、シート読み込み - ここから
    for i, s in enumerate(wb.sheetnames):

        # 機能シートなら機能シート読み込み関数、入力項目なら入力項目読み込み関数をそれぞれ実行。それ以外は読み込まない（他で読み込んでるので）
        if '入力項目' in s:
            d.inputItems[i] = read_sheet_input_items(wb, s)
        elif '【機能】' in s:
            d.functions[i] = read_sheet_function(wb[s])
        else:
            continue
    # 【機能】xxx、入力項目、シート読み込み - ここまで

    return d


def read_sheets_for_report(title, wb):
    d = ReportData()  # インスタンス化

    d.title = title  # タイトル設定

    d.history = read_sheet(wb, '改訂履歴')  # 改訂履歴

    d.layout = read_sheet_layout(wb['レイアウト'])    # レイアウト

    d.reportItems = read_sheet(wb, '帳票項目')    # 帳票項目読み込み

    d.events = read_sheet(wb, 'イベント一覧')    # イベント一覧読み込み

    return d


def read_sheets_for_mail(title, wb):
    d = MailData()  # インスタンス化

    d.title = title  # タイトル設定

    d.history = read_sheet(wb, '改訂履歴')    # 改訂履歴読み込み

    # メールテンプレート読み込み関数 : 関数化した意味 → 字下げかつエディタ上で閉じれるという自分用可読性だけ。使い回す予定なし。
    def read_mail_template(ws, dict):

        row = col = 1   # 行カーソルと列カーソルをセルの「A1」で初期化。ここから下に探索する。

        # 概要欄探索
        while ws.cell(row, col).value != '概要':
            row = row + 1
        dict['Overview'] = ws.cell(row + 1, col).value    # 概要欄ゲット

        # テンプレートの表を読み込み
        while ws.cell(row, col).value != '送信元(From)':    # 「送信元(From)」を探索
            row = row + 1

        # 送信元(From) のタイトルを固定で取得する。「 」「項目」「繰返」「備考」の4つ
        table_array = [[' ', '項目', '繰返', '備考']]

        # []探索用関数
        def search_func(s, pos_start):
            b = s.find(']', pos_start)
            a = s.rfind('[', pos_start, b)
            return a, b

        # 表全部を取得する
        for rowdata in ws.iter_rows(min_row=row, max_row=ws.max_row, min_col=1, max_col=4):  # 行の内容を指定の範囲で 1行ずつ取得。
            row_array = []
            # array.append([str(cell.value) if cell.value is not None else ' ' for cell in rowdata])

            for i, cell in enumerate(rowdata):    # 行データから1セルずつ取得

                if i == 1 and cell.value is not None:
                    # まず[]に挟まれた単語の一覧を取得する
                    sb_list = []  # 角カッコのSquareBrackets用配列
                    v = str(cell.value)  # v は cell.value の v
                    pos_start = 0
                    while True:

                        pos_start, pos_end = search_func(v, pos_start)

                        if pos_end == -1:
                            break
                        elif pos_start == -1:
                            pos_start = pos_end + 1
                            continue
                        else:
                            sb_list.append(v[pos_start + 1:pos_end])
                            pos_start = pos_end + 1

                    # その一覧を太字斜体に装飾する
                    for s in sb_list:
                        v = v.replace('[' + s + ']', '***[' + s + ']***')

                    row_array.append(v)
                else:
                    row_array.append(str(cell.value) if cell.value is not None else ' ')

            table_array.append(row_array)

        dict['MailTemplate'] = table_array

    read_mail_template(wb['メールテンプレート'], d.mailTemplate)    # メールテンプレート読み込み

    d.mailItems = read_sheet(wb, 'メール項目')    # 帳票項目読み込み

    # サンプル読み込み関数 : 関数化した意味 → 字下げかつエディタ上で閉じれるという自分用可読性だけ。使い回す予定なし。
    def read_mail_sample(ws, array):

        for row in range(1, ws.max_row + 1):
            array.append(str(ws.cell(row, 1).value) if ws.cell(row, 1).value is not None else '')

        array.append('\n')

    read_mail_sample(wb['サンプル'], d.sample)

    return d


def convert_thread(file):

    # エクセルブックを開いて、シートをクラスに読み込んで、配置用ディレクトリを作って、mdを生成する。
    # 読み込み関数は関数オブジェクトとして渡してもらう

    print(file)
    wb = openpyxl.load_workbook(file, data_only=True)  # ファイル読み込み

    # 各設計書用の読み込み関数を関数オブジェクトとして利用する。
    func = None
    # file_name = os.path.basename(file)
    if '画面設計書_' in file:
        func = read_sheets_for_screen_design
    elif '帳票設計書_' in file:
        func = read_sheets_for_report
    elif 'メール設計書_' in file:
        func = read_sheets_for_mail

    else:
        print('■ NoMatchFileName : ' + file)
        return

    d = None
    try:
        d = func(file[file.find('_') + 1: file.rfind('.')], wb)  # エクセルの各シート読み込み
    except shutil.Error:
        print('■ error : ' + file)
        return

    # アウトプット準備 : ディレクトリ作成
    dir_path = file.replace(file[:file.find('\\')], OUTPUT_DIRECTORY)    # 先頭のディレクトリパスを取得して、アウトプットディレクトリに置き換える
    file_name = os.path.splitext(os.path.basename(file))[0].rstrip()  # ファイル名（拡張子なし）
    dir_path = os.path.dirname(dir_path) + os.sep + file_name  # ↑のファイル名を付与したディレクトリにする。階層深くなるけどそういうもの。
    os.makedirs(dir_path + os.sep + 'img', exist_ok=True)    # img の階層まで一気にディレクトリ作成

    # css ファイル作成
    with open(dir_path + os.sep + 'style.css', mode='w', encoding='utf-8_sig') as f:
        f.write(CSS_TEXT)

    # md生成
    output_file_name = dir_path + os.sep + file_name + '.md'
    with open(output_file_name, mode='w', encoding='utf-8_sig') as f:
        f.write(d.generate_markdown() + '\n')


def exec():
    print('\n★★ 本処理 - start')

    # エクセルファイルの一覧を取得（or 指定）して順次読み込み。なんにせよリスト型になってればOK
    # 画面設計書
    # ls = glob.glob(WORK_DIRECTORY + '\\*画面設計書\\**\\*.xlsx', recursive=True)
    ls = ['work\\06.画面設計書\\本部管理\\商品値引管理\\画面設計書_SC07-13-01_クーポン変更.xlsx']

    # 帳票設計書
    # ls = glob.glob(WORK_DIRECTORY + '\\*帳票設計書\\**\\*.xlsx', recursive=True)
    # ls = ls + [
    #     'work\\15.帳票設計書\\店舗管理\\商品管理\\【機密(Ａ)】【新お届け】帳票設計書_FM19_チラシ商品Soldout表示リスト .xlsx',
    #     #            'work\\15.帳票設計書\\店舗管理\\精算管理\\【機密(Ａ)】【新お届け】帳票設計書_FM01_ネットスーパー売上集計表.xlsx',
    #     #            'work\\15.帳票設計書\\店舗管理\\集荷管理\\【機密(Ａ)】【新お届け】帳票設計書_FM02_お客様メモ.xlsx'
    # ]

    # メール設計書
    # ls = glob.glob(WORK_DIRECTORY + '\\*メール設計書\\**\\*.xlsx', recursive=True)
    # ls = ['work\\17.メール設計書\\スコープ管理\\メール設計書_ML08-003_衣料品番反映完了.xlsx', ]

    # ls = ls + [
    #     'work\\17.メール設計書\\スコープ管理\\メール設計書_ML08-004_アピール文言設定反映完了.xlsx',
    # ]

    # 作業開始
    with ThreadPoolExecutor(max_workers=6) as pool:
        pool.map(convert_thread, ls)

    print('\n★★ 本処理 - end')


def main():

    # 初期化 作業用ディレクトリにコピーしたり bakのお掃除とか
    if not IS_SKIP_INIT:  # スキップしない場合は init() 実施。あまり負荷ではないが、ワークディレクトリの削除とコピーを行っているので、ソースディレクトリが変わってないならやんなくてよい。プログラムの初実行とソースの設計書が変わったらやってね。
        init()

    # エクセル読み込む
    exec()


if __name__ == '__main__':

    print('★Start - PG')

    # カレントディレクトリをこのファイルが存在するところに
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    print('- カレントディレクトリ : ' + os.getcwd())

    # さぁがんばろ
    ORIGINAL_EXCEL_DIRECTORY = '20210930_エクセルをMDに'  # 元ネタが存在するディレクトリを指定。Dropbox上でもいいけど、容量節約で実態がないモードにしているとたぶん動かない。素直にローカルPCのWorkとかでお願いします。
    # ワークディレクトリのパスを生成
    WORK_DIRECTORY = 'work'
    # アウトプットディレクトリのパスを生成
    OUTPUT_DIRECTORY = 'zz_markdown_' + datetime.datetime.today().strftime("%Y%m%d%H%M%S")

    IS_SKIP_INIT = True   # 初回はかならずFalseで。2回目以降はめんどいからTrue（スキップする）でもよい。

    main()

    print('★End - PG')

# 以下自分用のメモ帳

# 006BA135 DBの緑色の色コード
